"""convert_excel_csv - 将 Excel 指定工作表转为 CSV。

依赖: openpyxl (pip install openpyxl)
渐进式约定: openpyxl 在 run() 内部局部 import。
"""
import csv
import os
from typing import Any, Dict

SKILL_META: Dict[str, Any] = {
    "name": "convert_excel_csv",
    "description": "将 Excel 指定工作表转为 CSV",
    "category": "document",
    "params": {
        "input_path": {
            "type": "str",
            "required": True,
            "description": "输入 Excel(.xlsx/.xls) 路径",
        },
        "output_path": {
            "type": "str",
            "required": False,
            "default": None,
            "description": "输出 CSV 路径；不填则按输入名生成",
        },
        "sheet_name": {
            "type": "str",
            "required": False,
            "default": None,
            "description": "工作表名；不填则使用活动工作表（第一个）",
        },
        "delimiter": {
            "type": "str",
            "required": False,
            "default": ",",
            "description": "CSV 分隔符，默认逗号",
        },
    },
    "dependencies": ["openpyxl"],
}


def run(**kwargs) -> Dict[str, Any]:
    """将 Excel 工作表导出为 CSV。返回输出路径、行数、列数。"""
    from openpyxl import load_workbook

    input_path = kwargs["input_path"]
    output_path = kwargs.get("output_path")
    sheet_name = kwargs.get("sheet_name")
    delimiter = kwargs.get("delimiter", ",")

    if not os.path.isfile(input_path):
        raise FileNotFoundError(f"输入 Excel 不存在: {input_path}")

    # 默认输出路径：输入名[_工作表名].csv
    if not output_path:
        base = os.path.splitext(input_path)[0]
        suffix = f"_{sheet_name}" if sheet_name else ""
        output_path = f"{base}{suffix}.csv"

    # read_only + data_only 提升读取性能并取计算后的值
    wb = load_workbook(input_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active

        rows = list(ws.iter_rows(values_only=True))
        with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
            # utf-8-sig (带 BOM) 让 Excel 直接打开不乱码
            writer = csv.writer(f, delimiter=delimiter)
            writer.writerows(rows)
    finally:
        wb.close()

    return {
        "output_path": output_path,
        "rows": len(rows),
        "columns": len(rows[0]) if rows else 0,
    }
