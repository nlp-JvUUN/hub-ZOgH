"""
Excel 格式加载器（.xlsx / .xls）

依赖：openpyxl（xlsx）、xlrd（xls）

处理策略：
  - 每个 sheet → 一个 table 块（markdown 表格）
  - 处理合并单元格：向所有组成单元格填充值
  - 限制每个 sheet 最多 5000 行、每行最多 100 列
  - sheet 名称记录在 section_path 中
"""

import logging
from pathlib import Path

from parsed_block_schema import ParsedDocument, ParsedBlock
from .base_loader import table_to_markdown

logger = logging.getLogger(__name__)

MAX_ROWS = 5000
MAX_COLS = 100


class XlsxLoader:
    """Excel 文件加载器。"""

    def supports(self, file_path: Path) -> bool:
        return file_path.suffix.lower() in (".xlsx", ".xls")

    def load(self, file_path: Path, meta: dict) -> ParsedDocument:
        logger.info(f"加载 Excel: {file_path.name}")

        if file_path.suffix.lower() == ".xls":
            return self._load_xls(file_path, meta)
        else:
            return self._load_xlsx(file_path, meta)

    def _load_xlsx(self, file_path: Path, meta: dict) -> ParsedDocument:
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "需要安装 openpyxl 来加载 .xlsx 文件:\n"
                "  pip install openpyxl"
            )

        wb = openpyxl.load_workbook(str(file_path), data_only=True)
        blocks = []

        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]

            # 先处理合并单元格：构建值映射
            merged_values = {}
            for merge_range in ws.merged_cells.ranges:
                top_left = ws.cell(merge_range.min_row, merge_range.min_col).value
                if top_left is not None:
                    for row in range(merge_range.min_row, merge_range.max_row + 1):
                        for col in range(merge_range.min_col, merge_range.max_col + 1):
                            merged_values[(row, col)] = top_left

            rows_data = []
            for row_idx, row in enumerate(ws.iter_rows(max_row=MAX_ROWS, max_col=MAX_COLS, values_only=False)):
                if row_idx >= MAX_ROWS:
                    rows_data.append([f"... (sheet 共 {ws.max_row}+ 行，仅显示前 {MAX_ROWS} 行)"])
                    break

                row_cells = []
                for col_idx, cell in enumerate(row):
                    if col_idx >= MAX_COLS:
                        break
                    # 合并单元格取值
                    pos = (cell.row, cell.column)
                    if pos in merged_values:
                        val = merged_values[pos]
                    else:
                        val = cell.value
                    row_cells.append(str(val) if val is not None else "")
                rows_data.append(row_cells)

            if rows_data:
                md_table = table_to_markdown(rows_data)
                blocks.append(ParsedBlock(
                    block_type="table",
                    content=md_table,
                    page_num=sheet_idx,
                    section_path=[f"Sheet: {sheet_name}"],
                ))

        wb.close()
        logger.info(f"  解析完成: {len(wb.sheetnames)} 个 sheet → {len(blocks)} 个表格块")
        return ParsedDocument(meta=meta, source=str(file_path), blocks=blocks)

    def _load_xls(self, file_path: Path, meta: dict) -> ParsedDocument:
        try:
            import xlrd
        except ImportError:
            raise ImportError(
                "需要安装 xlrd 来加载 .xls 文件:\n"
                "  pip install xlrd"
            )

        wb = xlrd.open_workbook(str(file_path))
        blocks = []

        for sheet_idx in range(wb.nsheets):
            ws = wb.sheet_by_index(sheet_idx)
            sheet_name = ws.name

            rows_data = []
            for row_idx in range(min(ws.nrows, MAX_ROWS)):
                row_cells = []
                for col_idx in range(min(ws.ncols, MAX_COLS)):
                    cell = ws.cell(row_idx, col_idx)
                    row_cells.append(str(cell.value) if cell.value != "" else "")
                rows_data.append(row_cells)

            if rows_data:
                md_table = table_to_markdown(rows_data)
                blocks.append(ParsedBlock(
                    block_type="table",
                    content=md_table,
                    page_num=sheet_idx,
                    section_path=[f"Sheet: {sheet_name}"],
                ))

        logger.info(f"  解析完成: {wb.nsheets} 个 sheet → {len(blocks)} 个表格块")
        return ParsedDocument(meta=meta, source=str(file_path), blocks=blocks)
